import openpyxl
from os import path

def load_workbook(wb_path):
    if path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    return openpyxl.Workbook()    

wb_path = 'test.xlsx'
wb = load_workbook(wb_path)

#đổi tên sheet
#sheet = wb['thinhit']
#sheet.title='Hobbies'
sheet = wb['Hobbies']
hobbies = [
    (1, 'Python'),
    (2, 'Autoit'),
    (3, 'automation'),
]
for hobby in hobbies:
    sheet.append(hobby)


wb.save(wb_path)
